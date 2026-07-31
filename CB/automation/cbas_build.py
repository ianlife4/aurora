# -*- coding: utf-8 -*-
"""
CBAS 報價彙整 — 本機建構腳本
讀取 4 家券商 xlsx + 元富 1150430 + 統一證附表 → 產出 JSON 並嵌入 HTML 模板。

執行：py -3.12 build.py
輸出：CBAS報價彙整.html  (與 cbas_data.json)
"""
import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl
import xlrd
import requests
import urllib3

# 櫃買中心憑證缺 Subject Key Identifier,新版 OpenSSL 會 CERTIFICATE_VERIFY_FAILED
# → 全站一律 verify=False (同 automation/fetch_twsa_bookbuilding.py 對證券商公會的處理)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

SRC_DIR = Path(__file__).parent
TEMPLATE_HTML = SRC_DIR / 'template.html'
OUT_HTML = SRC_DIR / 'CBAS報價彙整.html'
OUT_JSON = SRC_DIR / 'cbas_data.json'
CB_DB_PATH = Path(r'C:\Users\J.Chun\Desktop\總資料\CB\automation\cb_data.db')

BROKERS = ['富邦', '永豐金', '元大', '統一證', '群益', '台新元富']

# 手動補值 — 券商 xlsx 尚未收錄、但已在其 CBAS交易頁報價的 CB。
# 只在「該 broker 對該 cb_code 尚無報價」時併入(見 main());xlsx 一旦收錄就自動讓位、不重複。
# 收錄後可留著(去重會略過)或刪掉此項。
MANUAL_QUOTES = [
    # 65843 南俊國際三 — 統一證 CBAS交易頁手填 (2026-07-15)
    {
        'broker': '統一證', 'cb_code': '65843', 'cb_name': '南俊國際三',
        'tcri': 'TCRI6', 'premium_100': 7.88, 'discount_rate': 0.05,
        'expiration': '2028-06-22', 'conv_price': 685.0, 'cb_price': 140.0,
        'note': '手動補 · 統一證 CBAS交易 2026-07-15',
    },
]

FILE_PATTERNS = {
    '富邦':     re.compile(r'富邦.*\.xlsx?$', re.I),
    '永豐金':   re.compile(r'永豐金.*\.xlsx?$', re.I),
    '元大':     re.compile(r'^CBAS報價表.*\.xlsx?$', re.I),
    '統一證':   re.compile(r'統一證.*\.xlsx?$', re.I),
    '群益':     re.compile(r'群益.*\.xlsx?$', re.I),
    '台新元富': re.compile(r'台新.*option.*\.xlsx?$', re.I),
}
PRIMARY_PATTERN = re.compile(r'初級市場.*\.xlsx?$', re.I)

# 櫃買中心「轉(交)換公司債停止轉(交)換資訊」(取代元大已停供的同名分頁)
TPEX_BASE = 'https://www.tpex.org.tw'
TPEX_CBSUSPEND_API = TPEX_BASE + '/www/zh-tw/bond/cbSuspend'
TPEX_REFERER = TPEX_BASE + '/zh-tw/bond/announce/close.html'
TPEX_UA = ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
           '(KHTML, like Gecko) Chrome/120.0 Safari/537.36')

# 可轉債「新制」上路日: 2025-07-01 起【發行】之 CB 才可在股東會停止過戶期間請求轉換
# (不得變更股東名簿登記);此前發行者仍禁轉 = 元大備註的「舊制,停過期禁轉」。
# 新舊制只影響【股東會】事由;配股配息/現金增資 一律禁轉,新舊制皆同。
NEW_REGIME_DATE = '2025-07-01'


# ─── helpers ──────────────────────────────────────────

def normalize(s):
    if s is None:
        return ''
    return re.sub(r'\s+', '', str(s)).lower()


def find_col(headers, *keywords, exclude=None, all_required=False):
    """回傳 header 第一個符合的 col index（找不到回 -1）。"""
    excl = [normalize(e) for e in (exclude or [])]
    kws = [normalize(k) for k in keywords]
    for i, h in enumerate(headers):
        nh = normalize(h)
        if not nh:
            continue
        if any(e in nh for e in excl):
            continue
        if all_required:
            if all(k in nh for k in kws):
                return i
        else:
            if any(k in nh for k in kws):
                return i
    return -1


def find_header_row(rows, *required_keywords):
    """找第一個包含所有 required keywords 的 row。"""
    kws = [normalize(k) for k in required_keywords]
    for i, row in enumerate(rows):
        nrow = [normalize(v) for v in row]
        if all(any(k in c for c in nrow) for k in kws):
            return i, row
    return -1, None


def to_date_str(v):
    if v is None or v == '':
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    # Excel 空日期格 → time(0,0) → '00:00:00'。純時間值不是日期,當沒填。
    # (例: 統一證 xlsx 的 66931 廣閎科一 賣回日尚未填,整列 市價/賣回價 皆 0)
    if re.fullmatch(r'\d{1,2}:\d{2}(:\d{2})?', s):
        return None
    m = re.match(r'^(\d{4})[/-](\d{1,2})[/-](\d{1,2})', s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    # excel serial number?
    try:
        f = float(s)
        if 30000 < f < 80000:
            d = datetime(1899, 12, 30) + (datetime.fromtimestamp(0) - datetime.fromtimestamp(0))
            from datetime import timedelta
            return (datetime(1899, 12, 30) + timedelta(days=int(f))).strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        pass
    return s


def to_num(v):
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def to_ratio(v, thresh=0.5):
    """把「比率型」欄位統一成【小數】(0.032 = 3.2%)。

    🔴 各券商單位不一致 (2026-07-30 用戶抓到,30454 台灣大四):
        履約折現率  永豐金/元大/統一證 = 0.032 (小數)  vs  群益 3.0219、台新元富 3.25 (百分比數字)
        在外流通率  四家 = 0.9971 (小數)              vs  台新元富 99.71 (百分比,欄名還寫 (%))
      → 前端直接 ×100 顯示 → 群益變 302.19%、台新變 325.00% / 9971.43%,跟其他家差 100 倍。

    判準:比率型欄位的合理小數上限遠小於 1
      - 利率類 thresh=0.5:折現率不可能是 50%,所以 >0.5 一定是百分比數字 → /100
      - 比例類 thresh=1.5:流通比例最多 1.0 (=100%),>1.5 必是百分比 → /100
    """
    f = to_num(v)
    if f is None:
        return None
    return f / 100 if abs(f) > thresh else f


def to_parity(v, stock_price=None, conv_price=None):
    """轉換價值 parity,統一成【百元制】(股價/轉換價×100,100 = 剛好平價)。

    🔴 各券商三種尺度混用 (2026-07-30 用戶抓到):
        永豐金/統一證 = 97.80 / 100.0 (百元制)
        元大/群益/台新元富 = 0.9868 / 0.9602 / 1.0 (小數制)
      → 前端同表並列就差 100 倍。

    最可靠的解法不是猜單位,而是【自己算】:有股價+轉換價就直接推導 (精確且與券商無關);
    兩者缺一才退回檔案值,並用門檻 5 判尺度 (百元制 <5 等於股價不到轉換價 5%,近乎廢券,極罕見)。
    """
    sp, cp = to_num(stock_price), to_num(conv_price)
    if sp and cp:
        return sp / cp * 100
    f = to_num(v)
    if f is None:
        return None
    return f * 100 if abs(f) < 5 else f


def to_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def to_code(v):
    """CB 代號統一成字串（去掉小數點）。"""
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        if float(v) != int(v):
            return str(v)
        return str(int(v))
    s = str(v).strip()
    return s if s else None


def cell_val(row, idx):
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def load_xlsx(path, sheet_name):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]


def load_xls(path, sheet_name):
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_name(sheet_name)
    out = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            v = ws.cell_value(r, c)
            t = ws.cell_type(r, c)
            if t == 3 and v:  # date
                try:
                    tup = xlrd.xldate_as_tuple(v, wb.datemode)
                    v = datetime(*tup) if tup[0] else v
                except Exception:
                    pass
            row.append(v)
        out.append(row)
    return out


def _file_date_key(f):
    """從檔名抽日期當排序 key,沒抽到 fallback 到 mtime。最新的 win。"""
    m = re.search(r'(20\d{2})[\-_]?(\d{2})[\-_]?(\d{2})', f.name)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    # 民國日期 (1150622 = 115/06/22)
    m = re.search(r'(\d{3})(\d{2})(\d{2})', f.name)
    if m and 100 <= int(m.group(1)) <= 130:
        return (int(m.group(1)) + 1911, int(m.group(2)), int(m.group(3)))
    return (0, 0, 0)


def find_broker_files(src_dir):
    """每家券商挑「檔名日期最新」的那一份 (多版本共存時)。
    掃根目錄 + CB報/ 子資料夾 (gmail_fetch_cbas_all.py 把新檔丟到 CB報/)。"""
    files = list(src_dir.glob('*.xls*'))
    cb_sub = src_dir / 'CB報'
    if cb_sub.is_dir():
        files += list(cb_sub.glob('*.xls*'))
    result = {}
    for broker, pat in FILE_PATTERNS.items():
        cands = [f for f in files if pat.search(f.name)]
        if cands:
            cands.sort(key=_file_date_key, reverse=True)
            result[broker] = cands[0]
    return result


def extract_date_from_filename(filename):
    m = re.search(r'(20\d{2})-?(\d{1,2})-?(\d{2})', filename)
    if m:
        y, mo, d = m.group(1), m.group(2).zfill(2), m.group(3)
        if 1 <= int(mo) <= 12 and 1 <= int(d) <= 31:
            return f"{y}-{mo}-{d}"
    return None


# ─── parsers ──────────────────────────────────────────

def parse_fubon(path):
    rows = load_xlsx(path, 'CBOP報價表')
    hi, header = find_header_row(rows, '債券名稱', 'CB代號')
    if hi < 0:
        return []

    c = {
        'name':        find_col(header, '債券名稱'),
        'code':        find_col(header, 'CB代號'),
        'tcri':        find_col(header, '擔保', '評等', 'TCRI'),
        'premium100':  find_col(header, '百元報價', 'on100'),
        'premium_ref': find_col(header, '參考價'),
        'duration':    find_col(header, '期間', 'Duration'),
        'discount':    find_col(header, '折現率'),
        'cb_price':    find_col(header, 'CB市價', 'CBPx'),
        'parity':      find_col(header, '轉換價值', 'Parity'),
        'premium_pct': find_col(header, '折溢價'),
        'bond_floor':  find_col(header, '參考履約', 'BondFloor'),
        'expiration':  find_col(header, '選擇權到期', 'Expiration'),
        'put_date':    find_col(header, '賣回日', 'PutDate'),
        'put_price':   find_col(header, '賣回價', 'PutPx'),
        'conv_price':  find_col(header, 'CB轉換價', 'Conv.Px'),
        'stock_price': find_col(header, '股票市價', 'StockPx'),
        'outstanding': find_col(header, '流通餘額', 'Outstanding'),
    }

    bonds = []
    for r in rows[hi + 1:]:
        code = to_code(cell_val(r, c['code']))
        if not code:
            continue
        if not re.fullmatch(r'\d{3,7}', code):
            # 富邦 sheet 後面可能再接其他表，欄位錯位 → 停
            break
        bonds.append({
            'broker': '富邦',
            'cb_code': code,
            'cb_name': to_str(cell_val(r, c['name'])),
            'tcri': to_str(cell_val(r, c['tcri'])),
            'premium_100': to_num(cell_val(r, c['premium100'])),
            'premium_ref': to_num(cell_val(r, c['premium_ref'])),
            'duration': to_num(cell_val(r, c['duration'])),
            'discount_rate': to_ratio(cell_val(r, c['discount'])),
            'cb_price': to_num(cell_val(r, c['cb_price'])),
            'parity': to_parity(cell_val(r, c['parity']), cell_val(r, c['stock_price']), cell_val(r, c['conv_price'])),
            'premium_pct': to_num(cell_val(r, c['premium_pct'])),
            'bond_floor': to_num(cell_val(r, c['bond_floor'])),
            'expiration': to_date_str(cell_val(r, c['expiration'])),
            'put_date': to_date_str(cell_val(r, c['put_date'])),
            'put_price': to_num(cell_val(r, c['put_price'])),
            'conv_price': to_num(cell_val(r, c['conv_price'])),
            'stock_price': to_num(cell_val(r, c['stock_price'])),
            'outstanding_ratio': to_ratio(cell_val(r, c['outstanding']), 1.5),
        })
    return bonds


def parse_sinopac(path):
    rows = load_xls(path, '選擇權ASO報價')
    hi, header = find_header_row(rows, '代號', '選擇權到期日')
    if hi < 0:
        return []

    c = {
        'name':        find_col(header, '標的'),
        'code':        find_col(header, '代號'),
        'industry':    find_col(header, '產業'),
        'guarantee':   find_col(header, '是否有擔保', 'Guaranteed'),
        'tcri':        find_col(header, '評等', 'Rating'),
        'premium100':  find_col(header, '佰元'),
        'premium_ref': find_col(header, '市價參考'),
        'expiration':  find_col(header, '選擇權到期'),
        'put_date':    find_col(header, '賣回日'),
        'put_price':   find_col(header, '賣回價'),
        'discount':    find_col(header, '折現率', '履約利率'),
        'stock_price': find_col(header, '現股'),
        'conv_price':  find_col(header, '轉換價格'),
        'parity':      find_col(header, '轉換', '價值', all_required=True),
        'cb_price':    find_col(header, 'cb', '價格', all_required=True, exclude=['轉換']),
        'premium_pct': find_col(header, '折溢價'),
        'issue_size':  find_col(header, '發行金額'),
        'outstanding': find_col(header, '餘額比例'),
        'note':        find_col(header, '備註'),
    }

    bonds = []
    # row hi+1 是英文 sub-header, 跳過
    for r in rows[hi + 2:]:
        code = to_code(cell_val(r, c['code']))
        if not code:
            continue
        if not re.fullmatch(r'\d{3,7}', code):
            break
        bonds.append({
            'broker': '永豐金',
            'cb_code': code,
            'cb_name': to_str(cell_val(r, c['name'])),
            'industry': to_str(cell_val(r, c['industry'])),
            'guarantee': to_str(cell_val(r, c['guarantee'])),
            'tcri': to_str(cell_val(r, c['tcri'])),
            'premium_100': to_num(cell_val(r, c['premium100'])),
            'premium_ref': to_num(cell_val(r, c['premium_ref'])),
            'expiration': to_date_str(cell_val(r, c['expiration'])),
            'put_date': to_date_str(cell_val(r, c['put_date'])),
            'put_price': to_num(cell_val(r, c['put_price'])),
            'discount_rate': to_ratio(cell_val(r, c['discount'])),
            'stock_price': to_num(cell_val(r, c['stock_price'])),
            'conv_price': to_num(cell_val(r, c['conv_price'])),
            'parity': to_parity(cell_val(r, c['parity']), cell_val(r, c['stock_price']), cell_val(r, c['conv_price'])),
            'cb_price': to_num(cell_val(r, c['cb_price'])),
            'premium_pct': to_num(cell_val(r, c['premium_pct'])),
            'issue_size': to_num(cell_val(r, c['issue_size'])),
            'outstanding_ratio': to_ratio(cell_val(r, c['outstanding']), 1.5),
            'note': to_str(cell_val(r, c['note'])),
        })
    return bonds


def parse_yuanta(path):
    rows = load_xlsx(path, '金融交易部資產交換選擇權報價表')
    hi, header = find_header_row(rows, '代號', '百元報價')
    if hi < 0:
        return []

    c = {
        'name':        find_col(header, '名稱'),
        'code':        find_col(header, '代號'),
        'guarantee':   find_col(header, '擔保'),
        'tcri':        find_col(header, 'TCRI'),
        'premium100':  find_col(header, '百元報價'),
        'discount':    find_col(header, '折現率'),
        'expiration':  find_col(header, '選擇權到期'),
        'put_date':    find_col(header, '賣回日'),
        'duration':    find_col(header, '年期'),
        'put_price':   find_col(header, '賣回價'),
        'conv_price':  find_col(header, '轉換價', exclude=['值']),
        'parity':      find_col(header, '轉換價值'),
        'cb_price':    find_col(header, 'CB市價'),
        'premium_pct': find_col(header, '溢/折價', '溢折價', '折溢'),
        'unit_cost':   find_col(header, '參考單價'),
        'issue_size_lots': find_col(header, '發行張數'),
        'outstanding': find_col(header, '餘額', exclude=['比例']),
        'volatility':  find_col(header, '波動度'),
        'note':        find_col(header, '備註'),
    }

    bonds = []
    for r in rows[hi + 1:]:
        code = to_code(cell_val(r, c['code']))
        if not code:
            continue
        if not re.fullmatch(r'\d{3,7}', code):
            break
        # ★ 元大百元報價 ×100
        prem100 = to_num(cell_val(r, c['premium100']))
        if prem100 is not None:
            prem100 = round(prem100 * 100, 6)

        tcri_raw = cell_val(r, c['tcri'])
        tcri = f"TCRI {int(tcri_raw)}" if isinstance(tcri_raw, (int, float)) else (
            f"TCRI {tcri_raw}" if tcri_raw else None)

        bonds.append({
            'broker': '元大',
            'cb_code': code,
            'cb_name': to_str(cell_val(r, c['name'])),
            'guarantee': to_str(cell_val(r, c['guarantee'])),
            'tcri': tcri,
            'premium_100': prem100,
            'discount_rate': to_ratio(cell_val(r, c['discount'])),
            'expiration': to_date_str(cell_val(r, c['expiration'])),
            'put_date': to_date_str(cell_val(r, c['put_date'])),
            'duration': to_num(cell_val(r, c['duration'])),
            'put_price': to_num(cell_val(r, c['put_price'])),
            'conv_price': to_num(cell_val(r, c['conv_price'])),
            'parity': to_parity(cell_val(r, c['parity']), cell_val(r, c['stock_price']), cell_val(r, c['conv_price'])),
            'cb_price': to_num(cell_val(r, c['cb_price'])),
            'premium_pct': to_num(cell_val(r, c['premium_pct'])),
            'unit_cost': to_num(cell_val(r, c['unit_cost'])),
            'issue_size_lots': to_num(cell_val(r, c['issue_size_lots'])),
            'outstanding_ratio': to_ratio(cell_val(r, c['outstanding']), 1.5),
            'volatility_21d': to_num(cell_val(r, c['volatility'])),
            'note': to_str(cell_val(r, c['note'])),
        })
    return bonds


def parse_president(path):
    rows = load_xlsx(path, 'CBAS報價')
    hi, header = find_header_row(rows, 'CB代號', '百元報價')
    if hi < 0:
        return []

    c = {
        'name':        find_col(header, 'CB名稱'),
        'code':        find_col(header, 'CB代號'),
        'tcri':        find_col(header, '擔保', '評等'),
        'premium100':  find_col(header, '百元報價'),
        'discount':    find_col(header, '履約利率'),
        'expiration':  find_col(header, '選擇權到期'),
        'put_date':    find_col(header, '賣回日'),
        'put_price':   find_col(header, '賣回價'),
        'duration':    find_col(header, '年期'),
        'conv_price':  find_col(header, '轉換價格'),
        'stock_price': find_col(header, '標的股價'),
        'parity':      find_col(header, '轉換價值'),
        'cb_price':    find_col(header, 'CB市價'),
        'premium_pct': find_col(header, '折溢價'),
        'unit_cost':   find_col(header, '參考單張成本'),
        'issue_size_lots': find_col(header, '發行量'),
        'outstanding': find_col(header, '流通比例'),
        'note':        find_col(header, '注意事項'),
    }

    bonds = []
    for r in rows[hi + 1:]:
        code = to_code(cell_val(r, c['code']))
        if not code:
            continue
        # 統一證 sheet 末段是損益試算器，遇到非整數代號就停
        if not re.fullmatch(r'\d{3,7}', code):
            break
        bonds.append({
            'broker': '統一證',
            'cb_code': code,
            'cb_name': to_str(cell_val(r, c['name'])),
            'tcri': to_str(cell_val(r, c['tcri'])),
            'premium_100': to_num(cell_val(r, c['premium100'])),
            'discount_rate': to_ratio(cell_val(r, c['discount'])),
            'expiration': to_date_str(cell_val(r, c['expiration'])),
            'put_date': to_date_str(cell_val(r, c['put_date'])),
            'put_price': to_num(cell_val(r, c['put_price'])),
            'duration': to_num(cell_val(r, c['duration'])),
            'conv_price': to_num(cell_val(r, c['conv_price'])),
            'stock_price': to_num(cell_val(r, c['stock_price'])),
            'parity': to_parity(cell_val(r, c['parity']), cell_val(r, c['stock_price']), cell_val(r, c['conv_price'])),
            'cb_price': to_num(cell_val(r, c['cb_price'])),
            'premium_pct': to_num(cell_val(r, c['premium_pct'])),
            'unit_cost': to_num(cell_val(r, c['unit_cost'])),
            'issue_size_lots': to_num(cell_val(r, c['issue_size_lots'])),
            'outstanding_ratio': to_ratio(cell_val(r, c['outstanding']), 1.5),
            'note': to_str(cell_val(r, c['note'])),
        })
    return bonds


def parse_capital(path):
    """群益金鼎 - sheet '報價表'"""
    rows = load_xls(path, '報價表')
    hi, header = find_header_row(rows, '可轉債代號', '百元報價')
    if hi < 0:
        return []

    c = {
        'name':        find_col(header, '可轉債名稱'),
        'code':        find_col(header, '可轉債代號'),
        'premium100':  find_col(header, '百元報價'),
        'expiration':  find_col(header, '選擇權到期'),
        'duration':    find_col(header, '剩餘年限'),
        'put_date':    find_col(header, 'putday'),
        'tcri':        find_col(header, '評等', '擔保銀行'),
        'put_price':   find_col(header, '賣回價'),
        'discount':    find_col(header, '履約折現率'),
        'bond_floor':  find_col(header, '參考履約'),
        'conv_price':  find_col(header, '轉換價格'),
        'stock_price': find_col(header, '現股價格'),
        'cb_price':    find_col(header, '可轉債價格'),
        'parity':      find_col(header, 'parity'),
        # 🔴 find_col 預設是 any-match:原本 ('溢','折','價率') 會被第 8 欄「履約【折】現率」
        #    搶先命中 → 折溢價欄顯示的其實是履約利率 (2026-07-30 用戶抓到)。
        #    群益真正的欄名是「溢(折)價率」,用 '價率' 就夠精準且不會撞到「折現率」。
        'premium_pct': find_col(header, '價率', exclude=['折現']),
        'issue_size':  find_col(header, '發行總額'),
        'outstanding_lots': find_col(header, '餘額', exclude=['比例']),
        'outstanding': find_col(header, '流通比例'),
    }

    bonds = []
    for r in rows[hi + 1:]:
        code = to_code(cell_val(r, c['code']))
        if not code:
            continue
        if not re.fullmatch(r'\d{3,7}', code):
            break
        bonds.append({
            'broker': '群益',
            'cb_code': code,
            'cb_name': to_str(cell_val(r, c['name'])),
            'premium_100': to_num(cell_val(r, c['premium100'])),
            'expiration': to_date_str(cell_val(r, c['expiration'])),
            'duration': to_num(cell_val(r, c['duration'])),
            'put_date': to_date_str(cell_val(r, c['put_date'])),
            'tcri': to_str(cell_val(r, c['tcri'])),
            'put_price': to_num(cell_val(r, c['put_price'])),
            'discount_rate': to_ratio(cell_val(r, c['discount'])),
            'bond_floor': to_num(cell_val(r, c['bond_floor'])),
            'conv_price': to_num(cell_val(r, c['conv_price'])),
            'stock_price': to_num(cell_val(r, c['stock_price'])),
            'cb_price': to_num(cell_val(r, c['cb_price'])),
            'parity': to_parity(cell_val(r, c['parity']), cell_val(r, c['stock_price']), cell_val(r, c['conv_price'])),
            'premium_pct': to_num(cell_val(r, c['premium_pct'])),
            'issue_size': to_num(cell_val(r, c['issue_size'])),
            'issue_size_lots': to_num(cell_val(r, c['outstanding_lots'])),
            'outstanding_ratio': to_ratio(cell_val(r, c['outstanding']), 1.5),
        })
    return bonds


def parse_taishin(path):
    """台新元富(2025 台新證券 + 元富證券合併) - sheet '報價表'"""
    rows = load_xls(path, '報價表')
    hi, header = find_header_row(rows, 'CB代號', '百元')
    if hi < 0:
        return []

    c = {
        'name':        find_col(header, 'CB', exclude=['代號', '參考價']),
        'code':        find_col(header, 'CB代號'),
        'premium100':  find_col(header, '百元價'),
        'duration':    find_col(header, '剩餘'),
        'expiration':  find_col(header, '選擇權到期'),
        'put_date':    find_col(header, 'putday'),
        'premium_ref': find_col(header, '選擇權參考價'),
        'tcri':        find_col(header, '評等', '擔保行'),
        'put_price':   find_col(header, '賣回價格'),
        'discount':    find_col(header, '折現率'),
        'bond_floor':  find_col(header, '參考履約'),
        'conv_price':  find_col(header, '轉換價', exclude=['參考']),
        'stock_price': find_col(header, '現股', '參考價', all_required=True),
        'cb_price':    find_col(header, '可轉債', '參考價', all_required=True),
        'parity':      find_col(header, 'parity'),
        'premium_pct': find_col(header, '折溢價'),
        'outstanding_lots': find_col(header, '流通', '餘額', all_required=True),
        'outstanding': find_col(header, '在外', '流通率', all_required=True),
        'industry':    find_col(header, '產業別'),
    }

    bonds = []
    for r in rows[hi + 1:]:
        code = to_code(cell_val(r, c['code']))
        if not code:
            continue
        if not re.fullmatch(r'\d{3,7}', code):
            break
        bonds.append({
            'broker': '台新元富',
            'cb_code': code,
            'cb_name': to_str(cell_val(r, c['name'])),
            'premium_100': to_num(cell_val(r, c['premium100'])),
            'duration': to_num(cell_val(r, c['duration'])),
            'expiration': to_date_str(cell_val(r, c['expiration'])),
            'put_date': to_date_str(cell_val(r, c['put_date'])),
            'premium_ref': to_num(cell_val(r, c['premium_ref'])),
            'tcri': to_str(cell_val(r, c['tcri'])),
            'put_price': to_num(cell_val(r, c['put_price'])),
            'discount_rate': to_ratio(cell_val(r, c['discount'])),
            'bond_floor': to_num(cell_val(r, c['bond_floor'])),
            'conv_price': to_num(cell_val(r, c['conv_price'])),
            'stock_price': to_num(cell_val(r, c['stock_price'])),
            'cb_price': to_num(cell_val(r, c['cb_price'])),
            'parity': to_parity(cell_val(r, c['parity']), cell_val(r, c['stock_price']), cell_val(r, c['conv_price'])),
            'premium_pct': to_num(cell_val(r, c['premium_pct'])),
            'issue_size_lots': to_num(cell_val(r, c['outstanding_lots'])),
            'outstanding_ratio': to_ratio(cell_val(r, c['outstanding']), 1.5),
            'industry': to_str(cell_val(r, c['industry'])),
        })
    return bonds


def parse_capital_primary(path):
    """群益金鼎 - sheet '初級市場' → 補入 primary_market.in_pipeline / .announced"""
    try:
        rows = load_xls(path, '初級市場')
    except Exception:
        return []
    hi, header = find_header_row(rows, '可轉債代號', '掛牌日')
    if hi < 0:
        return []
    c = {
        'code':        find_col(header, '可轉債代號'),
        'name':        find_col(header, '可轉債名稱'),
        'duration':    find_col(header, '發行年期'),
        'amount':      find_col(header, '發行金額'),
        'put_terms':   find_col(header, '發行條件', '賣回'),
        'conv_price':  find_col(header, '轉換價'),
        'underwriter': find_col(header, '主辦', '承銷商'),
        'tcri':        find_col(header, '信用評等', '擔保'),
        'list_date':   find_col(header, '掛牌日'),
        'cbas_date':   find_col(header, 'CBAS承作日'),
        'note':        find_col(header, '備註'),
    }
    out = []
    for r in rows[hi + 1:]:
        code = to_code(cell_val(r, c['code']))
        if not code or not re.fullmatch(r'\d{3,7}', code):
            continue
        list_date_raw = cell_val(r, c['list_date'])
        list_date_str = str(list_date_raw or '')
        is_pipeline = '詢圈' in list_date_str or '競拍' in list_date_str
        dur = cell_val(r, c['duration'])
        out.append({
            'code': code,
            'name': to_str(cell_val(r, c['name'])),
            'tcri': re.sub(r'^TCRI\s*', '', to_str(cell_val(r, c['tcri'])) or ''),
            'method': '詢圈' if '詢圈' in list_date_str else ('競拍' if '競拍' in list_date_str else '掛牌'),
            'amount_yi': to_num(cell_val(r, c['amount'])),
            'duration_y': (str(int(dur)) + '年') if isinstance(dur, (int, float)) else to_str(dur),
            'price': '群益',  # 暫存承銷商在 price，下方會更新
            'underwriter': to_str(cell_val(r, c['underwriter'])),
            'put_terms': to_str(cell_val(r, c['put_terms'])),
            'conv_price': None,
            'stock_price': None,
            'premium_pct_str': to_str(cell_val(r, c['conv_price'])),  # 含 (溢價%)
            'inquiry_period': list_date_str if is_pipeline else None,
            'list_date': to_date_str(list_date_raw) if not is_pipeline else None,
            'cbas_date': to_date_str(cell_val(r, c['cbas_date'])),
            '_is_pipeline': is_pipeline,
        })
    return out


def fetch_close_conversion_tpex(timeout=30):
    """櫃買中心官方「轉(交)換公司債停止轉(交)換資訊」— 每日更新,涵蓋全市場。

    為何是權威來源: 台灣 CB 不分上市/上櫃一律在櫃買中心交易,所以這一支就是全市場
    (今日樣本含 140201 遠東新、15601 中砂 等上市公司 CB),不需要另抓證交所。
    元大 xlsx 那張同名分頁本來就是抄這份;元大 2026-07 起精簡檔案不再提供 → 改直抓官方。

    流程: cbSuspend API 回「每日檔案索引」(日期新→舊) → 取最新一日的 CSV → cp950 解碼
          → 取 BODY 列 (欄: 債券代碼/債券簡稱/停止起日/停止迄日/事由)。
    回 (rows, data_date);任何失敗回 ([], None) 讓 build 繼續走 fallback。
    """
    try:
        s = requests.Session()
        s.headers.update({'User-Agent': TPEX_UA, 'Referer': TPEX_REFERER})
        s.verify = False   # 櫃買憑證缺 SKI,見檔頭 urllib3 註解
        idx = s.get(TPEX_CBSUSPEND_API, timeout=timeout).json()
        rows = idx.get('tables', [{}])[0].get('data', [])
        if not rows:
            return [], None
        data_date, csv_path = rows[0][0], rows[0][2]   # rows[0] = 最新一日; 欄2 = CSV

        raw = s.get(TPEX_BASE + csv_path, timeout=timeout).content
        text = None
        for enc in ('cp950', 'big5', 'utf-8-sig', 'utf-8'):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            return [], data_date

        out = []
        for line in text.splitlines():
            if not line.startswith('BODY,'):
                continue
            cells = re.findall(r'"([^"]*)"', line)
            if len(cells) < 5:
                continue
            code = cells[0].strip()
            if not re.fullmatch(r'\d{5,6}', code):   # CB 代號 5-6 碼 (140201 等 6 碼也要)
                continue
            out.append({
                'code': code,
                'name': cells[1].strip(),
                'start_date': _tpex_date(cells[2]),
                'end_date': _tpex_date(cells[3]),
                'reason': cells[4].strip(),
            })
        return out, data_date
    except Exception as e:
        print(f'  [WARN] 櫃買中心停止轉換抓取失敗: {e}')
        return [], None


def _tpex_date(s):
    """2026/06/22 → 2026-06-22"""
    m = re.match(r'(\d{4})/(\d{1,2})/(\d{1,2})', (s or '').strip())
    return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}' if m else ''


def parse_close_conversion(path):
    """[fallback] 元大 xlsx 的停止轉換分頁。主來源已改 fetch_close_conversion_tpex(),
    這只在櫃買中心抓不到時墊檔;元大 2026-07 起多半已無此分頁。"""
    try:
        rows = load_xlsx(path, '轉(交)換公司債停止轉(交)換資訊')
    except KeyError:
        # 元大精簡版 xlsx (如 CB報/CBAS報價表20260703.xlsx) 只留主報價分頁 → 回空,不要 crash 整個 build
        print('  [WARN] 元大檔缺「停止轉換」分頁,close_conversion 略過')
        return []
    hi, header = find_header_row(rows, '債券代碼', '停止')
    if hi < 0:
        return []
    c = {
        'code':   find_col(header, '債券代碼'),
        'name':   find_col(header, '債券簡稱'),
        'start':  find_col(header, '起日'),
        'end':    find_col(header, '迄日'),
        'reason': find_col(header, '事由'),
    }
    out = []
    for r in rows[hi + 1:]:
        code = to_code(cell_val(r, c['code']))
        if not code:
            continue
        out.append({
            'code': code,
            'name': to_str(cell_val(r, c['name'])),
            'start_date': to_date_str(cell_val(r, c['start'])),
            'end_date': to_date_str(cell_val(r, c['end'])),
            'reason': to_str(cell_val(r, c['reason'])),
        })
    return out


def parse_aso_close(path):
    try:
        rows = load_xlsx(path, 'ASO到期標的停止轉換資訊')
    except KeyError:
        # 元大 6 月後 xlsx 拿掉這個 sheet → 回空就好,不要 crash 整個 build
        return []
    hi, header = find_header_row(rows, '標的代號', '選擇權到期日')
    if hi < 0:
        return []
    c = {
        'expiration': find_col(header, '選擇權到期日'),
        'code':       find_col(header, '標的代號'),
        'name':       find_col(header, '標的名稱'),
        'status':     find_col(header, '狀態'),
        'period':     find_col(header, '停止轉換期間'),
        'reason':     find_col(header, '事由'),
        'note':       find_col(header, '備註'),
    }
    out = []
    for r in rows[hi + 1:]:
        code = to_code(cell_val(r, c['code']))
        if not code:
            continue
        out.append({
            'expiration': to_date_str(cell_val(r, c['expiration'])),
            'code': code,
            'name': to_str(cell_val(r, c['name'])),
            'status': to_str(cell_val(r, c['status'])),
            'period': to_str(cell_val(r, c['period'])),
            'reason': to_str(cell_val(r, c['reason'])),
            'note': to_str(cell_val(r, c['note'])),
        })
    return out


def parse_upcoming_maturity(path):
    """讀 統一證 xlsx「近期賣回.到期.贖回之CB」sheet。
    含兩個 section: 「最近一個月可賣回之CB」(1m) + 「最近三個月到期之CB」(3m)。
    回傳 list[{code, name, cb_price, put_price, put_date, put_yield_pct,
              outstanding_lots, outstanding_ratio, section}]。
    """
    SHEET = '近期賣回.到期.贖回之CB'
    try:
        rows = load_xlsx(path, SHEET)
    except KeyError:
        return []

    section_starts = []
    for i, r in enumerate(rows):
        v = str(r[0] or '').strip()
        if '最近一個月可賣回' in v:
            section_starts.append((i, '1m'))
        elif '最近三個月到期' in v:
            section_starts.append((i, '3m'))

    by_code = {}  # dedupe: 同檔同 put_date → 保留 section='1m' (更急迫的)
    for k, (sec_start, label) in enumerate(section_starts):
        sec_end = section_starts[k+1][0] if k+1 < len(section_starts) else len(rows)
        hi, header = -1, None
        for j in range(sec_start, sec_end):
            head_str = ' '.join(str(v or '') for v in rows[j][:8])
            if '債券代號' in head_str and ('賣回日期' in head_str or '到期日' in head_str):
                hi, header = j, rows[j]
                break
        if hi < 0:
            continue
        c = {
            'code':       find_col(header, '債券代號'),
            'name':       find_col(header, '標的債券'),
            'cb_price':   find_col(header, 'CB市價'),
            'put_price':  find_col(header, '賣回價'),
            'put_date':   find_col(header, '賣回日期', '到期日'),
            'yield_pct':  find_col(header, '賣回收益率'),
            'lots':       find_col(header, '流通餘額'),
            'ratio':      find_col(header, '餘額比例'),
        }
        for r in rows[hi + 1:sec_end]:
            code = to_code(cell_val(r, c['code']))
            if not code:
                break
            item = {
                'code': code,
                'name': to_str(cell_val(r, c['name'])),
                'cb_price': to_num(cell_val(r, c['cb_price'])),
                'put_price': to_num(cell_val(r, c['put_price'])),
                'put_date': to_date_str(cell_val(r, c['put_date'])),
                'put_yield_pct': to_num(cell_val(r, c['yield_pct'])),
                'outstanding_lots': to_num(cell_val(r, c['lots'])),
                'outstanding_ratio': to_num(cell_val(r, c['ratio'])),
                'section': label,
            }
            key = code
            if key not in by_code:
                by_code[key] = item
            elif label == '1m':
                # 1m 比 3m 更急 → 覆蓋
                by_code[key] = item
    return list(by_code.values())


def join_conv_price_from_db(items):
    """從 cb_data.db (CB 競拍管理 DB) issued 表補 conv_price。"""
    if not items or not CB_DB_PATH.exists():
        return items
    try:
        import sqlite3
        conn = sqlite3.connect(str(CB_DB_PATH))
        codes = list({it['code'] for it in items})
        ph = ','.join('?' for _ in codes)
        mp = {row[0]: row[1] for row in conn.execute(
            f'SELECT cb_code, conv_price FROM issued WHERE cb_code IN ({ph})', codes
        )}
        conn.close()
        for it in items:
            it['conv_price'] = mp.get(it['code'])
        n = sum(1 for it in items if it.get('conv_price') is not None)
        print(f'  conv_price join: {n}/{len(items)} 命中')
    except Exception as e:
        print(f'  [WARN] conv_price join 失敗: {e}')
    return items


def join_stock_price_from_bonds(items, bonds):
    """從 merge_bonds 後的 bond 物件補 stock_price (各券商報價快照)。
    用來算 stock_price < conv_price → CB OTM 訊號。"""
    if not items or not bonds:
        return items
    bond_by_code = {b.get('cb_code'): b for b in bonds}
    n = 0
    for it in items:
        b = bond_by_code.get(it.get('code'))
        sp = b.get('stock_price') if b else None
        it['stock_price'] = sp
        if sp is not None:
            n += 1
    print(f'  stock_price join: {n}/{len(items)} 命中')
    return items


def _pm_method(inquiry_text):
    """從『詢圈/競拍 期間』文字判方式。"""
    t = inquiry_text or ''
    if '競拍' in t:
        return '競拍'
    if '詢圈' in t:
        return '詢圈'
    return None


def parse_primary_market_v2(path):
    """2026 起 富邦/元富 新版『CB初級市場資訊』sheet。三段:
       詢圈/競拍 + 送件 → in_pipeline;董事會通過 → announced。
       (比舊 masterlink 版完整,涵蓋每檔 pipeline CB。)"""
    # 新 sheet max_row 常是百萬幽靈列 → 用 read_only iter,連續 40 空列就停
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb['CB初級市場資訊']
    rows = []
    blank = 0
    for r in ws.iter_rows(values_only=True):
        if all(c is None or str(c).strip() == '' for c in r[:10]):
            blank += 1
            if blank >= 40:
                break
            rows.append(list(r))
            continue
        blank = 0
        rows.append(list(r))
    wb.close()
    pipeline, announced = [], []
    # 找三段 section header row index
    sec = {}
    for i, r in enumerate(rows[:200]):
        v = str(r[0] or '').strip()
        if '詢圈' in v and '競拍' in v and '標的' in v:
            sec['inquiry'] = i
        elif v == '送件標的' or ('送件' in v and len(v) < 8):
            sec['filed'] = i
        elif '董事會' in v and ('通過' in v or '發行' in v):
            sec['board'] = i
    # Layout A: 詢圈/競拍 + 送件 (col 0-9)
    def parse_layout_a(start, end):
        out = []
        for r in rows[start:end]:
            code = to_code(cell_val(r, 0))
            if not code or not code[0].isdigit() or len(code) < 5:
                continue
            dur = cell_val(r, 2)
            inquiry = to_str(cell_val(r, 8))
            out.append({
                'code': code,
                'name': to_str(cell_val(r, 1)),
                'tcri': to_str(cell_val(r, 7)),
                'method': _pm_method(inquiry),
                'amount_yi': to_num(cell_val(r, 3)),
                'duration_y': (str(int(dur)) + '年') if isinstance(dur, (int, float)) else to_str(dur),
                'price': None,
                'underwriter': to_str(cell_val(r, 6)),
                'premium_pct_str': to_str(cell_val(r, 5)),
                'conv_price': to_num(cell_val(r, 5)),
                'stock_price': None,
                'put_terms': to_str(cell_val(r, 4)),
                'inquiry_period': inquiry,
                'list_date': to_date_str(cell_val(r, 9)),
                'cbas_date': None,
            })
        return out
    # Layout B: 董事會通過 (col 0-9: code,name,期間,金額,董事會日期,承銷商,TCRI,方式,產業,股價)
    def parse_layout_b(start, end):
        out = []
        for r in rows[start:end]:
            code = to_code(cell_val(r, 0))
            if not code or not code[0].isdigit() or len(code) < 5:
                continue
            dur = cell_val(r, 2)
            method_txt = to_str(cell_val(r, 7))
            out.append({
                'code': code,
                'name': to_str(cell_val(r, 1)),
                'tcri': to_str(cell_val(r, 6)),
                'method': _pm_method(method_txt),
                'amount_yi': to_num(cell_val(r, 3)),
                'duration_y': (str(int(dur)) + '年') if isinstance(dur, (int, float)) else to_str(dur),
                'board_date': to_date_str(cell_val(r, 4)),
                'underwriter': to_str(cell_val(r, 5)),
                'industry': to_str(cell_val(r, 8)),
                'stock_price': to_num(cell_val(r, 9)),
            })
        return out

    idx_sorted = sorted(v for v in sec.values())
    def next_sec(after):
        nxt = [i for i in idx_sorted if i > after]
        return min(nxt) if nxt else len(rows)

    if 'inquiry' in sec:
        pipeline += parse_layout_a(sec['inquiry'] + 2, next_sec(sec['inquiry']))
    if 'filed' in sec:
        pipeline += parse_layout_a(sec['filed'] + 2, next_sec(sec['filed']))
    if 'board' in sec:
        announced += parse_layout_b(sec['board'] + 2, len(rows))

    return {'in_pipeline': pipeline, 'announced': announced}


def parse_primary_market(path):
    # dispatcher: 新版 sheet 'CB初級市場資訊' → v2;舊 masterlink '工作表1' → 舊邏輯
    try:
        import openpyxl as _ox
        names = _ox.load_workbook(path, read_only=True).sheetnames
    except Exception:
        names = []
    if 'CB初級市場資訊' in names:
        return parse_primary_market_v2(path)
    rows = load_xlsx(path, '工作表1')

    pipe_section = None
    ann_section = None
    for i, r in enumerate(rows):
        v = str(r[0] or '')
        if pipe_section is None and ('送件' in v or '詢圈標' in v):
            pipe_section = i
        if '董事會' in v and '公告' in v:
            ann_section = i

    pipeline, announced = [], []

    def find_first_data_header(start_idx, end_idx):
        for j in range(start_idx, min(end_idx, len(rows))):
            r = rows[j]
            head_str = ''.join(str(v or '') for v in r[:5])
            if ('代號' in head_str or '代碼' in head_str):
                return j, r
        return -1, None

    if pipe_section is not None:
        end = ann_section if ann_section else len(rows)
        hi, header = find_first_data_header(pipe_section + 1, end)
        if hi >= 0:
            c = {
                'code':        find_col(header, '標的', '代號', all_required=True),
                'name':        find_col(header, '標的', '名稱', all_required=True),
                'tcri':        find_col(header, 'TCRI'),
                'method':      find_col(header, '承銷', '方式', all_required=True),
                'amount':      find_col(header, '金額', '億元'),
                'duration':    find_col(header, '發行', '期間', all_required=True),
                'price':       find_col(header, '發行', '價格', all_required=True),
                'underwriter': find_col(header, '承銷券商'),
                'premium_pct': find_col(header, '溢價率'),
                'conv_price':  find_col(header, '轉換價', exclude=['值']),
                'put_terms':   find_col(header, '賣回價', '年/'),
                'inquiry':     find_col(header, '詢圈', '競拍'),
                'list_date':   find_col(header, '掛牌日'),
                'cbas_date':   find_col(header, '拆解日'),
                'stock_price': find_col(header, '股價'),
            }
            for r in rows[hi + 1:end]:
                code = to_code(cell_val(r, c['code']))
                if not code or not code[0].isdigit():
                    continue
                dur = cell_val(r, c['duration'])
                pipeline.append({
                    'code': code,
                    'name': to_str(cell_val(r, c['name'])),
                    'tcri': to_str(cell_val(r, c['tcri'])),
                    'method': to_str(cell_val(r, c['method'])),
                    'amount_yi': to_num(cell_val(r, c['amount'])),
                    'duration_y': (str(int(dur)) + '年') if isinstance(dur, (int, float)) else to_str(dur),
                    'price': to_str(cell_val(r, c['price'])),
                    'underwriter': to_str(cell_val(r, c['underwriter'])),
                    'premium_pct_str': to_str(cell_val(r, c['premium_pct'])),
                    'conv_price': to_num(cell_val(r, c['conv_price'])),
                    'stock_price': to_num(cell_val(r, c['stock_price'])),
                    'put_terms': to_str(cell_val(r, c['put_terms'])),
                    'inquiry_period': to_str(cell_val(r, c['inquiry'])),
                    'list_date': to_date_str(cell_val(r, c['list_date'])),
                    'cbas_date': to_date_str(cell_val(r, c['cbas_date'])),
                })

    if ann_section is not None:
        hi, header = find_first_data_header(ann_section + 1, len(rows))
        if hi >= 0:
            c = {
                'code':        find_col(header, '代碼', '標的代號'),
                'name':        find_col(header, '標的名稱', '名稱'),
                'tcri':        find_col(header, 'TCRI'),
                'method':      find_col(header, '承銷', '方式', all_required=True),
                'amount':      find_col(header, '金額'),
                'duration':    find_col(header, '年限', '期間'),
                'price':       find_col(header, '發行', '價格', all_required=True),
                'underwriter': find_col(header, '承銷券商'),
            }
            for r in rows[hi + 1:]:
                code = to_code(cell_val(r, c['code']))
                if not code or not code[0].isdigit():
                    continue
                dur = cell_val(r, c['duration'])
                announced.append({
                    'code': code,
                    'name': to_str(cell_val(r, c['name'])),
                    'tcri': to_str(cell_val(r, c['tcri'])),
                    'method': to_str(cell_val(r, c['method'])),
                    'amount_yi': to_num(cell_val(r, c['amount'])),
                    'duration_y': (str(int(dur)) + '年') if isinstance(dur, (int, float)) else to_str(dur),
                    'price': to_str(cell_val(r, c['price'])),
                    'underwriter': to_str(cell_val(r, c['underwriter'])),
                })

    return {'in_pipeline': pipeline, 'announced': announced}


# ─── merge & build ──────────────────────────────────────

def _listing_dates_from_db(codes):
    """cb_data.db issued 表取 listing_date (判新舊制用)。"""
    if not codes or not CB_DB_PATH.exists():
        return {}
    try:
        import sqlite3
        conn = sqlite3.connect(str(CB_DB_PATH))
        ph = ','.join('?' for _ in codes)
        mp = {r[0]: r[1] for r in conn.execute(
            f'SELECT cb_code, listing_date FROM issued WHERE cb_code IN ({ph})', list(codes))}
        conn.close()
        return mp
    except Exception as e:
        print(f'  [WARN] listing_date join 失敗: {e}')
        return {}


def _regime_note(sc, listing_date):
    """新舊制備註。只有【股東會】事由才有新舊制差異 (配股配息/現增 一律禁轉,不標)。
    以掛牌日當發行日 proxy 判 NEW_REGIME_DATE 分界 (CB 發行日與掛牌日僅差數日)。"""
    if '股東' not in (sc.get('reason') or ''):
        return None
    if not listing_date:
        return None
    return ('新制,停過期可轉(不得變更股東名簿)' if listing_date[:10] >= NEW_REGIME_DATE
            else '舊制,停過期禁轉')


def build_aso_close(upcoming_mat, close_conv):
    """重建「ASO 到期停轉」表 (元大 2026-07 起停供該分頁)。

    原表在講的事: 選擇權快到期/快賣回的 CB,若標的正在停止轉換 → 可能轉不掉。
    元大是把兩份資料對起來而已,所以用同樣組合重建:
        近期賣回/到期清單 (統一證 xlsx parse_upcoming_maturity — 仍有供)
          LEFT JOIN 櫃買中心停止轉換 (fetch_close_conversion_tpex)
    欄位對應原表: 選擇權到期日=賣回/到期日、狀態=1m→Next put / 3m→到期、
                 停止轉換期間+事由=櫃買、備註=新舊制。
    LEFT JOIN: 沒對到停轉的也留著 (原表亦然,period/reason 留空)。
    """
    if not upcoming_mat:
        return []
    stop_by = {r['code']: r for r in close_conv}
    listing = _listing_dates_from_db([m['code'] for m in upcoming_mat])
    out = []
    for m in upcoming_mat:
        sc = stop_by.get(m['code'])
        out.append({
            'expiration': m.get('put_date'),
            'code': m['code'],
            'name': m.get('name'),
            'status': '到期' if m.get('section') == '3m' else 'Next put',
            'period': f"{sc['start_date']}~{sc['end_date']}" if sc else None,
            'reason': sc['reason'] if sc else None,
            'note': _regime_note(sc, listing.get(m['code'])) if sc else None,
        })
    out.sort(key=lambda x: (x.get('expiration') or '9999-99-99'))
    return out


def merge_bonds(all_quotes):
    by_code = {}
    for q in all_quotes:
        code = q['cb_code']
        if not code:
            continue
        b = by_code.get(code)
        if not b:
            b = by_code[code] = {
                'cb_code': code,
                'cb_name': q.get('cb_name'),
                'cb_name_alt': [],
                'tcri': None,
                'industry': None,
                'guarantee': None,
                'expiration': None,
                'put_date': None,
                'put_price': None,
                'conv_price': None,
                'cb_price': None,
                'stock_price': None,
                'parity': None,
                'brokers': [],
                'quotes': [],
            }
        # 補缺 metadata（先到的優先）
        for k in ['cb_name', 'tcri', 'industry', 'guarantee', 'expiration',
                  'put_date', 'put_price', 'conv_price', 'cb_price',
                  'stock_price', 'parity']:
            if b.get(k) is None and q.get(k) is not None:
                b[k] = q.get(k)
        # alt names
        nm = q.get('cb_name')
        if nm and nm != b['cb_name'] and nm not in b['cb_name_alt']:
            b['cb_name_alt'].append(nm)
        b['brokers'].append(q['broker'])
        b['quotes'].append({k: v for k, v in q.items() if k != 'cb_code'})

    return sorted(by_code.values(), key=lambda x: x['cb_code'])


def main():
    print('=' * 60)
    print('CBAS 報價彙整 build')
    print('=' * 60)
    files = find_broker_files(SRC_DIR)
    for k, v in files.items():
        print(f'  {k}: {v.name}')

    sources, broker_counts = {}, {}
    all_quotes = []

    parsers = {
        '富邦':     parse_fubon,
        '永豐金':   parse_sinopac,
        '元大':     parse_yuanta,
        '統一證':   parse_president,
        '群益':     parse_capital,
        '台新元富': parse_taishin,
    }

    for broker, parser in parsers.items():
        if broker not in files:
            print(f'  [WARN] no file for {broker}')
            continue
        f = files[broker]
        try:
            quotes = parser(f)
        except Exception as e:
            print(f'  [ERR] {broker}: {e}')
            quotes = []
        all_quotes.extend(quotes)
        broker_counts[broker] = len(quotes)
        sources[broker] = {'file': f.name, 'date': extract_date_from_filename(f.name)}
        print(f'  {broker}: {len(quotes)} quotes')

    # 手動補值:只在該 broker 對該 cb_code 尚無報價時併入(xlsx 收錄後自動讓位)
    _have = {(q.get('broker'), q.get('cb_code')) for q in all_quotes}
    for mq in MANUAL_QUOTES:
        if (mq['broker'], mq['cb_code']) not in _have:
            all_quotes.append(dict(mq))
            broker_counts[mq['broker']] = broker_counts.get(mq['broker'], 0) + 1
            print(f"  [MANUAL] {mq['broker']} {mq['cb_code']} {mq.get('cb_name', '')}")

    bonds = merge_bonds(all_quotes)

    # 停止轉換 — 主來源改櫃買中心官方 (每日更新、全市場);元大檔僅在網路失敗時墊檔
    close_conv, cc_date = fetch_close_conversion_tpex()
    if close_conv:
        print(f'  停止轉換 ← 櫃買中心 {cc_date}: {len(close_conv)} 筆')
    else:
        close_conv = parse_close_conversion(files['元大']) if '元大' in files else []
        print(f'  [WARN] 櫃買中心抓不到,回退元大檔: {len(close_conv)} 筆')

    upcoming_mat = parse_upcoming_maturity(files['統一證']) if '統一證' in files else []
    upcoming_mat = join_conv_price_from_db(upcoming_mat)
    upcoming_mat = join_stock_price_from_bonds(upcoming_mat, bonds)

    # ASO 到期停轉 — 元大停供該分頁 → 用 近期賣回/到期 × 櫃買停轉 重建
    aso_close = build_aso_close(upcoming_mat, close_conv)
    _aso_hit = sum(1 for x in aso_close if x.get('period'))
    print(f'  ASO 到期停轉 ← 重建: {len(aso_close)} 筆 (其中 {_aso_hit} 筆對到停轉期間)')

    # 元富初級市場 — 2026 起改夾在富邦 mail (CB初級市場資訊YYYYMMDD.xlsx)。掃根目錄 + CB報/,取最新。
    pm_candidates = [f for f in SRC_DIR.glob('*.xlsx') if PRIMARY_PATTERN.search(f.name)]
    _pm_sub = SRC_DIR / 'CB報'
    if _pm_sub.is_dir():
        pm_candidates += [f for f in _pm_sub.glob('*.xlsx') if PRIMARY_PATTERN.search(f.name)]
    pm_files = sorted(pm_candidates, key=_file_date_key, reverse=True)
    primary_market = parse_primary_market(pm_files[0]) if pm_files else {'in_pipeline': [], 'announced': []}
    if pm_files:
        sources['元富'] = {'file': pm_files[0].name, 'date': extract_date_from_filename(pm_files[0].name)}

    # 群益的初級市場合併進來（用 cb_code 去重，已存在的不覆蓋）
    if '群益' in files:
        cap_pm = parse_capital_primary(files['群益'])
        existing = {p['code'] for p in primary_market['in_pipeline']} | {p['code'] for p in primary_market['announced']}
        for p in cap_pm:
            if p['code'] in existing:
                continue
            is_pipe = p.pop('_is_pipeline')
            (primary_market['in_pipeline'] if is_pipe else primary_market['announced']).append(p)

    print(f'\n→ {len(bonds)} bonds | {len(close_conv)} close | {len(aso_close)} aso | {len(upcoming_mat)} upcoming_maturity')
    print(f'→ pipeline {len(primary_market["in_pipeline"])} | announced {len(primary_market["announced"])}')

    data = {
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'sources': sources,
        'broker_counts': broker_counts,
        'bonds': bonds,
        'primary_market': primary_market,
        'close_conversion': close_conv,
        'close_conversion_date': cc_date,   # 櫃買中心資料日期 (民國 115/07/16);回退元大檔時為 None
        'aso_close_conversion': aso_close,
        'upcoming_maturity': upcoming_mat,
    }

    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'\n[OK] JSON → {OUT_JSON}')

    if not TEMPLATE_HTML.exists():
        print(f'[ERR] template not found: {TEMPLATE_HTML}')
        sys.exit(1)
    template = TEMPLATE_HTML.read_text(encoding='utf-8')
    json_str = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
    pat = re.compile(r'(<script id="DATA"[^>]*>)(.*?)(</script>)', re.DOTALL)
    new_html = pat.sub(lambda m: m.group(1) + json_str + m.group(3), template, count=1)
    OUT_HTML.write_text(new_html, encoding='utf-8')
    print(f'[OK] HTML → {OUT_HTML}')


if __name__ == '__main__':
    main()
