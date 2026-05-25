#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 更新器
負責將解析後的資料寫入 CB競拍整理 AAAS.xlsb.xlsx
  - 已發行1（Sheet index 8）：來自元富 Gmail
  - CB競拍結果（Sheet index 9）：來自 TWSE
"""
import shutil
import warnings
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

from config import EXCEL_PATH, SHEET_INDEX, ISSUED_COL, AUCTION_COL

warnings.filterwarnings('ignore')


# ── 共用工具 ──────────────────────────────────────────────────────────────────

def _load_existing_codes(excel_path: Path, sheet_idx: int, code_col: int = 0) -> set[str]:
    """讀取指定 Sheet 的現有 CB 代碼（去除重複）"""
    try:
        sheets = pd.read_excel(str(excel_path), sheet_name=None, header=0)
        df = list(sheets.values())[sheet_idx]
        codes = df.iloc[:, code_col].dropna().astype(str)
        return set(codes.str.strip().str.replace(r'\.0$', '', regex=True))
    except Exception as e:
        print(f'  ⚠️  讀取現有代碼失敗：{e}')
        return set()


def _backup_excel(excel_path: Path) -> Path:
    """備份原始 Excel"""
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup = excel_path.with_name(f'{excel_path.stem}_backup_{ts}{excel_path.suffix}')
    shutil.copy2(str(excel_path), str(backup))
    print(f'  備份：{backup.name}')
    return backup


def _get_sheet_by_index(wb, idx: int):
    """用 index 取 Sheet（因工作表名稱是中文）"""
    return wb.worksheets[idx]


def _last_data_row(ws) -> int:
    """找最後一個有資料的列號（1-based）"""
    last = ws.max_row
    while last > 1 and all(ws.cell(last, c).value is None for c in range(1, 5)):
        last -= 1
    return last


def _to_float(val) -> float | None:
    if val is None or val == '':
        return None
    try:
        return float(str(val).replace(',', '').strip())
    except (ValueError, TypeError):
        return None


# ── 已發行1 更新（來自元富 Gmail）────────────────────────────────────────────

def update_issued_sheet(records: list[dict], dry_run: bool = False) -> int:
    """
    將元富解析結果寫入「已發行1」工作表。
    新 cb_code → 新增列；既有 cb_code → 補齊空欄位（不覆寫已填的）。
    回傳：新增筆數（不含補齊）
    """
    if not records:
        print('已發行1：無新資料')
        return 0

    # 先 dedupe records — 同 cb_code 多筆 (masterlink + unisec) 用 merge_records()
    # 避免後面 INSERT 重複 row,以及 _merge_into_existing_issued 後蓋掉好欄位
    records = _merge_records_by_cb(records)

    existing = _load_existing_codes(EXCEL_PATH, SHEET_INDEX['已發行1'], code_col=0)

    new_recs    = [r for r in records if str(r.get('cb_code','')).replace('.0','') not in existing]
    existing_recs = [r for r in records if str(r.get('cb_code','')).replace('.0','') in existing]

    # 對既有列做空欄位補齊（最常見：conv_price 後來才公告）
    if existing_recs and not dry_run:
        merged = _merge_into_existing_issued(existing_recs)
        if merged:
            print(f'已發行1：補齊 {merged} 筆既有列的空欄位')

    if not new_recs:
        if not existing_recs:
            print(f'已發行1：{len(records)} 筆均已存在，無需新增')
        return 0

    print(f'已發行1：待新增 {len(new_recs)} 筆')
    for r in new_recs:
        print(f"  + {r['cb_code']} {r.get('company','')}  TCRI:{r.get('tcri','')}  "
              f"{r.get('amount','')}億  {r.get('term','')}  "
              f"掛牌:{r.get('listing_date','')}  發行價:{r.get('issue_price_raw','')}")

    if dry_run:
        print('  （dry-run，未寫入）')
        return len(new_recs)

    _backup_excel(EXCEL_PATH)
    wb = load_workbook(str(EXCEL_PATH))
    ws = _get_sheet_by_index(wb, SHEET_INDEX['已發行1'])
    last_row = _last_data_row(ws)

    for i, r in enumerate(new_recs):
        row = last_row + 1 + i
        c = ISSUED_COL

        # 代碼（5碼）
        try:
            ws.cell(row, c['cb_code']+1).value = int(r['cb_code'])
        except ValueError:
            ws.cell(row, c['cb_code']+1).value = r['cb_code']

        ws.cell(row, c['company']+1).value     = r.get('company', '')
        ws.cell(row, c['tcri']+1).value        = r.get('tcri', '')
        ws.cell(row, c['method']+1).value      = r.get('method', '競拍')
        ws.cell(row, c['amount']+1).value      = r.get('amount')
        ws.cell(row, c['term']+1).value        = r.get('term', '')
        ws.cell(row, c['underwriter']+1).value = r.get('underwriter', '')

        # 發行價格（原始字串，可能是 "100~103" 或 "104.95"）
        ws.cell(row, c['issue_price']+1).value  = r.get('issue_price_raw', '') or r.get('issue_price')
        ws.cell(row, c['conv_price']+1).value   = r.get('conv_price')
        ws.cell(row, c['conv_price2']+1).value  = r.get('conv_value', '') or r.get('premium_rate', '')
        ws.cell(row, c['put_cond']+1).value     = r.get('put_cond', '')
        ws.cell(row, c['receipt_date']+1).value = r.get('receipt_date', '')
        ws.cell(row, c['eff_date']+1).value     = r.get('eff_date', '')
        ws.cell(row, c['bid_period']+1).value   = r.get('bid_period', '')
        ws.cell(row, c['listing_date']+1).value = r.get('listing_date', '')
        ws.cell(row, c['capital']+1).value      = r.get('capital')

        # 股票代碼（前4碼）
        stock = r.get('stock_code', '')
        try:
            ws.cell(row, c['stock_code']+1).value = int(stock)
        except (ValueError, TypeError):
            ws.cell(row, c['stock_code']+1).value = stock

    wb.save(str(EXCEL_PATH))
    print(f'  已發行1：成功新增 {len(new_recs)} 筆')
    return len(new_recs)


def _merge_records_by_cb(records: list[dict]) -> list[dict]:
    """同 cb_code 多筆 record (如 masterlink + unisec 都有 17172) 用「欄位 merge」:
    先到的 record 為主,後到的只補空欄位。回傳 dedupe 後的 list。

    為什麼必要: 若用「後蓋前」(dict comprehension),unisec record conv_price=None
    會覆蓋 masterlink record conv_price=76,接著 update merge 階段又因 new_v=None 跳過
    → conv_price 永遠寫不進 DB → 競拍建議價算不出。實證: 17172/34913/47491/76101 4 筆中招。
    """
    by_code: dict[str, dict] = {}
    for r in records:
        code = str(r.get('cb_code', '')).replace('.0', '')
        if not code:
            continue
        if code not in by_code:
            by_code[code] = dict(r)
        else:
            base = by_code[code]
            for k, v in r.items():
                old = base.get(k)
                if (old is None or (isinstance(old, str) and not old.strip())
                        or (isinstance(old, str) and old.strip() in ('未定','未公告','-','待定'))
                        ) and v not in (None, '', 'nan'):
                    base[k] = v
    return list(by_code.values())


def _merge_into_existing_issued(records: list[dict]) -> int:
    """對「已發行1」既有列補齊空欄位（不覆寫已填內容）。
    回傳實際更新筆數。"""
    if not records:
        return 0
    code_idx = ISSUED_COL['cb_code']

    # records 已 dedupe (caller update_issued_sheet 已跑 _merge_records_by_cb)
    by_code = {str(r.get('cb_code','')).replace('.0',''): r for r in records}

    _backup_excel(EXCEL_PATH)
    wb = load_workbook(str(EXCEL_PATH))
    ws = _get_sheet_by_index(wb, SHEET_INDEX['已發行1'])
    last_row = _last_data_row(ws)

    # 對映表：record key → ISSUED_COL key
    field_pairs = [
        ('company',      'company'),
        ('tcri',         'tcri'),
        ('method',       'method'),
        ('amount',       'amount'),
        ('term',         'term'),
        ('underwriter',  'underwriter'),
        ('issue_price_raw','issue_price'),
        ('conv_price',   'conv_price'),
        ('conv_value',   'conv_price2'),
        ('premium_rate', 'conv_price2'),
        ('put_cond',     'put_cond'),
        ('receipt_date', 'receipt_date'),
        ('eff_date',     'eff_date'),
        ('bid_period',   'bid_period'),
        ('listing_date', 'listing_date'),
        ('capital',      'capital'),
    ]

    n_updated = 0
    for row in range(2, last_row + 1):
        cell_v = ws.cell(row, code_idx + 1).value
        if cell_v is None:
            continue
        cb = str(cell_v).replace('.0', '').strip()
        if cb not in by_code:
            continue
        r = by_code[cb]
        changed = False
        for src_key, col_key in field_pairs:
            if src_key not in r:
                continue
            new_v = r.get(src_key)
            if new_v is None or (isinstance(new_v, str) and not new_v.strip()):
                continue
            col = ISSUED_COL.get(col_key)
            if col is None:
                continue
            old_v = ws.cell(row, col + 1).value
            old_str = str(old_v).strip() if old_v is not None else ''
            is_empty = (old_v is None
                        or (isinstance(old_v, str) and not old_v.strip())
                        or old_str in ('未定', '未公告', '-', '待定', 'nan', 'NaT', 'None')
                        or (isinstance(old_v, (int, float)) and old_v == 0
                            and col_key in ('conv_price','amount','capital')))
            if is_empty:
                ws.cell(row, col + 1).value = new_v
                changed = True
        if changed:
            n_updated += 1

    if n_updated:
        wb.save(str(EXCEL_PATH))
    return n_updated


# ── CB競拍結果 更新（來自 TWSE）─────────────────────────────────────────────

def update_auction_sheet(twse_records: list[dict], industry_map: dict = None,
                         dry_run: bool = False) -> int:
    # industry_map 參數保留但不寫入（個股資料由使用者自行維護）
    """
    將 TWSE 競拍結果寫入「CB競拍結果」工作表
    twse_records: twse_scraper.clean_twse_record 回傳的清單
    industry_map: {stock_code: {main_biz, sub_biz, biz_desc}} 來自個股 sheet
    回傳：新增筆數
    """
    if not twse_records:
        print('CB競拍結果：無新資料')
        return 0

    existing = _load_existing_codes(EXCEL_PATH, SHEET_INDEX['CB競拍結果'], code_col=4)  # CB代號欄

    new_recs = [r for r in twse_records
                if str(r.get('cb_code', '')).replace('.0', '') not in existing]

    if not new_recs:
        print(f'CB競拍結果：{len(twse_records)} 筆均已存在，無需更新')
        return 0

    print(f'CB競拍結果：待新增 {len(new_recs)} 筆')
    for r in new_recs:
        print(f"  + {r['cb_code']} {r.get('sec_name','')}  "
              f"開標:{r.get('open_date','')}  均價:{r.get('weighted_avg','?')}  "
              f"承銷價:{r.get('actual_price','?')}")

    if dry_run:
        print('  （dry-run，未寫入）')
        return len(new_recs)

    # 讀取已有的最大序號
    existing_max_serial = _get_max_serial(SHEET_INDEX['CB競拍結果'])

    wb = load_workbook(str(EXCEL_PATH))
    ws = _get_sheet_by_index(wb, SHEET_INDEX['CB競拍結果'])
    last_row = _last_data_row(ws)

    industry_map = industry_map or {}

    for i, r in enumerate(new_recs):
        row = last_row + 1 + i
        c   = AUCTION_COL
        serial = existing_max_serial + i + 1

        stock = r.get('stock_code', '')

        ws.cell(row, c['serial']+1).value        = serial
        ws.cell(row, c['auction_date']+1).value  = r.get('open_date', '')
        ws.cell(row, c['company']+1).value       = r.get('sec_name', '')
        ws.cell(row, c['stock_code']+1).value    = _to_int(stock)
        ws.cell(row, c['cb_code']+1).value       = _to_int(r.get('cb_code', ''))
        ws.cell(row, c['issue_amount']+1).value  = r.get('auction_amount_yi')
        ws.cell(row, c['auction_amount']+1).value= r.get('auction_amount_yi')
        ws.cell(row, c['guarantee']+1).value     = r.get('guarantee', '')
        ws.cell(row, c['auction_lots']+1).value  = r.get('auction_lots')
        ws.cell(row, c['min_bid_pct']+1).value   = r.get('min_bid_price')
        ws.cell(row, c['bid_date']+1).value      = r.get('bid_start', '')
        ws.cell(row, c['weighted_avg']+1).value  = r.get('weighted_avg')
        ws.cell(row, c['min_award_pct']+1).value = r.get('min_award')
        ws.cell(row, c['max_award_pct']+1).value = r.get('max_award')
        ws.cell(row, c['avg_award_pct']+1).value = r.get('weighted_avg')
        ws.cell(row, c['listing_date']+1).value  = r.get('transfer_date', '')
        ws.cell(row, c['lead_mgr']+1).value      = r.get('lead_mgr', '')
        ws.cell(row, c['total_award_amt']+1).value = r.get('total_award')
        ws.cell(row, c['total_valid']+1).value   = r.get('total_valid')
        ws.cell(row, c['valid_lots']+1).value    = r.get('valid_lots')
        ws.cell(row, c['actual_price']+1).value  = r.get('actual_price')

    wb.save(str(EXCEL_PATH))
    print(f'  CB競拍結果：成功新增 {len(new_recs)} 筆')
    return len(new_recs)


def _get_max_serial(sheet_idx: int) -> int:
    """取得 CB競拍結果 現有最大序號"""
    try:
        sheets = pd.read_excel(str(EXCEL_PATH), sheet_name=None, header=0)
        df = list(sheets.values())[sheet_idx]
        col = df.iloc[:, 0].dropna()
        nums = pd.to_numeric(col, errors='coerce').dropna()
        return int(nums.max()) if len(nums) > 0 else 0
    except Exception:
        return 0


def _to_int(val) -> int | str:
    try:
        return int(str(val).replace('.0', ''))
    except (ValueError, TypeError):
        return val


# ── 讀取個股產業資料 ──────────────────────────────────────────────────────────

def load_industry_map() -> dict[str, dict]:
    """
    從個股工作表（sheet index 11）讀取產業資料
    回傳 {stock_code_str: {main_biz, sub_biz, biz_desc, capital}}
    """
    try:
        sheets = pd.read_excel(str(EXCEL_PATH), sheet_name=None, header=0)
        df = list(sheets.values())[SHEET_INDEX['個股']]
        # 欄位：代碼, 股名, 主業, 副業, 主業描述, 市值, 指數, 其他, KY
        result = {}
        for _, row in df.iterrows():
            code = str(row.iloc[0]).replace('.0', '').strip()
            if not code or code == 'nan':
                continue
            result[code] = {
                'company':  str(row.iloc[1]) if pd.notna(row.iloc[1]) else '',
                'main_biz': str(row.iloc[2]) if pd.notna(row.iloc[2]) else '',
                'sub_biz':  str(row.iloc[3]) if pd.notna(row.iloc[3]) else '',
                'biz_desc': str(row.iloc[4]) if pd.notna(row.iloc[4]) else '',
                'capital':  row.iloc[5]      if pd.notna(row.iloc[5]) else None,
            }
        print(f'  載入個股產業資料：{len(result)} 筆')
        return result
    except Exception as e:
        print(f'  ⚠️  讀取個股 sheet 失敗：{e}')
        return {}
